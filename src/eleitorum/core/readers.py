"""Per-format file readers for EleitorUM.

Handles XLSX, XLSM, XLS, ODS, CSV, and TSV input files.

Design trade-off note: ReadResult.rows is ``list[tuple[Any, ...]]`` rather than
a lazy iterator. Materializing into a list is intentional:
- The largest expected input is ~150k rows (~50 MB peak RAM per RESEARCH.md PERF
  benchmark); this is well within a typical desktop machine's headroom.
- A lazy iterator would require keeping the file handle and the openpyxl
  read-only workbook open past the function boundary, complicating resource
  cleanup and making ``wb.close()`` semantics error-prone.
- detection.py and pipeline.py both need random access to determine header row
  position; a list is the natural data structure for that.

Security notes:
- T-1-02-02: every file-opening call is wrapped in try/except PermissionError
  → re-raised as FileAccessError (INP-13).
- T-1-02-03: openpyxl uses read_only=True, data_only=True — streaming mode,
  no full workbook in-memory load (PERF-03).
- T-1-02-05: extension whitelist check happens before any I/O; unknown
  extension raises UnsupportedFormatError without touching the file contents.
"""

from __future__ import annotations

import csv
import dataclasses
import pathlib
from typing import Any

import openpyxl  # type: ignore[import-untyped]
import pandas as pd  # type: ignore[import-untyped]
import xlrd  # type: ignore[import-untyped]

from eleitorum.core.errors import FileAccessError, UnsupportedFormatError

# ---------------------------------------------------------------------------
# Module-level constants
# ---------------------------------------------------------------------------

SUPPORTED_EXTENSIONS: frozenset[str] = frozenset(
    {".xlsx", ".xlsm", ".xls", ".ods", ".csv", ".tsv"}
)

# First 64KB of a CSV/TSV file is sampled for encoding detection (INP-07).
_ENCODING_SAMPLE_BYTES: int = 64 * 1024


# ---------------------------------------------------------------------------
# Data classes
# ---------------------------------------------------------------------------


@dataclasses.dataclass(frozen=True)
class SheetInfo:
    """Metadata about a single sheet in a multi-sheet Excel/ODS workbook.

    ``approximate_row_count`` is derived from ``ws.max_row`` in openpyxl
    read-only mode; per RESEARCH.md Pitfall 2 this value can be stale (the
    XLSX may have been saved with cached dimensions). It is used only for
    display in the Phase 2 sheet-picker dialog — never as a hard row count.

    ``is_empty`` is True when the sheet has at most one non-empty row (i.e. a
    header with no data rows, or a completely blank sheet).
    """

    name: str
    approximate_row_count: int  # from ws.max_row; approximate only
    is_empty: bool  # True if sheet has ≤ 1 non-empty row (header only or no data)


@dataclasses.dataclass
class ReadResult:
    """The result of reading any supported input file.

    rows:
        All non-trailing-empty rows as a list of tuples of raw cell values.
        Consumer modules normalise types; readers.py does no coercion.
    sheet_name:
        The name of the sheet read (Excel/ODS), or None for CSV/TSV.
    skipped_trailing_empty:
        Count of all-empty rows that were stripped from the tail.
    raw_bytes_sample:
        First ``_ENCODING_SAMPLE_BYTES`` bytes of the file for encoding
        detection; None for binary formats (XLSX/XLS/ODS).
    """

    rows: list[tuple[Any, ...]]
    sheet_name: str | None
    skipped_trailing_empty: int
    raw_bytes_sample: bytes | None


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    """Return True if every cell in ``row`` is None or blank after stripping."""
    return all(cell is None or str(cell).strip() == "" for cell in row)


def _strip_trailing_empty(
    rows: list[tuple[Any, ...]],
) -> tuple[list[tuple[Any, ...]], int]:
    """Strip all-empty rows from the tail of ``rows``.

    Returns the surviving rows list and the count of dropped rows.
    """
    if not rows:
        return [], 0
    count = 0
    while rows and _is_empty_row(rows[-1]):
        rows.pop()
        count += 1
    return rows, count


# ---------------------------------------------------------------------------
# Per-format readers
# ---------------------------------------------------------------------------


def read_xlsx(path: pathlib.Path, sheet_name: str | None = None) -> ReadResult:
    """Read an XLSX or XLSM file using openpyxl in read-only streaming mode.

    ``openpyxl.load_workbook(path, read_only=True, data_only=True)`` is
    non-negotiable: this is the PERF-03 contract that prevents loading the
    entire workbook into RAM on 150k-row files.

    Security note (T-1-02-02): PermissionError and FileNotFoundError are both
    wrapped as FileAccessError so the caller sees a consistent PT-PT message.
    """
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except PermissionError:
        raise FileAccessError(path=path, mode="read")
    except FileNotFoundError:
        raise FileAccessError(path=path, mode="read")

    chosen_sheet: str = sheet_name or wb.sheetnames[0]
    ws = wb[chosen_sheet]

    raw: list[tuple[Any, ...]] = list(ws.iter_rows(values_only=True))
    wb.close()

    stripped, skipped_count = _strip_trailing_empty(raw)
    return ReadResult(
        rows=stripped,
        sheet_name=chosen_sheet,
        skipped_trailing_empty=skipped_count,
        raw_bytes_sample=None,
    )


def read_xls(path: pathlib.Path, sheet_name: str | None = None) -> ReadResult:
    """Read a legacy XLS file using xlrd (XLS-only; xlrd 2.x does not read XLSX).

    Uses ``on_demand=True`` for memory-friendliness — sheets are loaded only
    when accessed.

    Note: xlrd cells carry typed ``.value`` attributes. For numeric cells,
    ``cell.value`` returns a Python float; TRF-02 in transform.py handles the
    ``14891.0 → "14891"`` conversion downstream.
    """
    try:
        wb_xls = xlrd.open_workbook(str(path), on_demand=True)
    except PermissionError:
        raise FileAccessError(path=path, mode="read")
    except FileNotFoundError:
        raise FileAccessError(path=path, mode="read")

    if sheet_name is not None:
        sheet = wb_xls.sheet_by_name(sheet_name)
        chosen_sheet: str = sheet_name
    else:
        sheet = wb_xls.sheet_by_index(0)
        chosen_sheet = wb_xls.sheet_names()[0]

    raw: list[tuple[Any, ...]] = [
        tuple(cell.value for cell in row) for row in sheet.get_rows()
    ]
    wb_xls.release_resources()

    stripped, skipped_count = _strip_trailing_empty(raw)
    return ReadResult(
        rows=stripped,
        sheet_name=chosen_sheet,
        skipped_trailing_empty=skipped_count,
        raw_bytes_sample=None,
    )


def read_ods(path: pathlib.Path, sheet_name: str | None = None) -> ReadResult:
    """Read an ODS file using pandas + odfpy (``engine="odf"``).

    The header row is prepended as the first tuple because downstream
    detection.py scores each row to find the actual header position and
    expects the raw file layout including that header row.

    Note: pandas 3.0 Copy-on-Write is now default — no explicit copy needed.
    """
    try:
        sheet_idx: str | int = sheet_name if sheet_name is not None else 0
        df = pd.read_excel(path, engine="odf", sheet_name=sheet_idx, header=0)
    except PermissionError:
        raise FileAccessError(path=path, mode="read")
    except FileNotFoundError:
        raise FileAccessError(path=path, mode="read")

    # Determine the resolved sheet name
    if sheet_name is not None:
        chosen_sheet: str | None = sheet_name
    else:
        # pandas doesn't expose the sheet name when reading by index; use ExcelFile
        try:
            with pd.ExcelFile(path, engine="odf") as xf:
                chosen_sheet = xf.sheet_names[0] if xf.sheet_names else None
        except Exception:
            chosen_sheet = None

    # Prepend header row then data rows
    header_tuple: tuple[Any, ...] = tuple(df.columns.tolist())
    data_rows: list[tuple[Any, ...]] = [
        tuple(row) for row in df.itertuples(index=False, name=None)
    ]
    all_rows: list[tuple[Any, ...]] = [header_tuple] + data_rows

    stripped, skipped_count = _strip_trailing_empty(all_rows)
    return ReadResult(
        rows=stripped,
        sheet_name=chosen_sheet,
        skipped_trailing_empty=skipped_count,
        raw_bytes_sample=None,
    )


def read_csv_like(
    path: pathlib.Path,
    delimiter: str = ",",
    encoding: str = "utf-8-sig",
) -> ReadResult:
    """Read a delimited text file (CSV or TSV) using stdlib csv.

    Opens the file binary-first to capture ``raw_bytes_sample`` for the
    encoding-detection handoff to detection.py (plan 03). Then re-opens in
    text mode with the caller-supplied encoding.

    The ``encoding`` parameter defaults to ``"utf-8-sig"`` (UTF-8 with BOM
    strip). The pipeline.py orchestrator in plan 06 will wire detection.py's
    output into this parameter for actual runtime use.
    """
    try:
        with open(path, mode="rb") as fb:
            sample = fb.read(_ENCODING_SAMPLE_BYTES)
        with open(path, mode="r", encoding=encoding, newline="") as ft:
            reader = csv.reader(ft, delimiter=delimiter)
            rows: list[tuple[Any, ...]] = [tuple(row) for row in reader]
    except PermissionError:
        raise FileAccessError(path=path, mode="read")
    except FileNotFoundError:
        raise FileAccessError(path=path, mode="read")

    stripped, skipped_count = _strip_trailing_empty(rows)
    return ReadResult(
        rows=stripped,
        sheet_name=None,
        skipped_trailing_empty=skipped_count,
        raw_bytes_sample=sample,
    )


# ---------------------------------------------------------------------------
# Multi-sheet metadata
# ---------------------------------------------------------------------------


def _count_nonempty_rows_xlsx(ws: Any, max_check: int = 5) -> int:
    """Count non-empty rows in the first ``max_check`` rows of an openpyxl worksheet."""
    count = 0
    for row in ws.iter_rows(values_only=True, max_row=max_check):
        if any(cell is not None and str(cell).strip() != "" for cell in row):
            count += 1
    return count


def list_sheets(path: pathlib.Path) -> list[SheetInfo]:
    """Return per-sheet metadata for multi-sheet Excel/ODS files.

    Returns ``[]`` for CSV/TSV (no sheet concept).

    The ``approximate_row_count`` field is derived from ``ws.max_row`` in
    openpyxl read-only mode (RESEARCH.md Pitfall 2: this value can be stale).
    It is used only for display in the Phase 2 sheet-picker dialog.

    ``is_empty`` is determined by checking the first 5 rows (per RESEARCH.md
    Pitfall 8 — Alunos-style header-only sheets): if fewer than 2 non-empty
    rows are found, the sheet is flagged as empty.
    """
    ext = path.suffix.lower()

    if ext in {".csv", ".tsv"}:
        return []

    if ext in {".xlsx", ".xlsm"}:
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except PermissionError:
            raise FileAccessError(path=path, mode="read")
        except FileNotFoundError:
            raise FileAccessError(path=path, mode="read")

        result: list[SheetInfo] = []
        for name in wb.sheetnames:
            ws = wb[name]
            approx_rows: int = ws.max_row or 0
            nonempty = _count_nonempty_rows_xlsx(ws, max_check=5)
            is_empty = nonempty < 2
            result.append(SheetInfo(name=name, approximate_row_count=approx_rows, is_empty=is_empty))
        wb.close()
        return result

    if ext == ".xls":
        try:
            wb_xls = xlrd.open_workbook(str(path), on_demand=True)
        except PermissionError:
            raise FileAccessError(path=path, mode="read")
        except FileNotFoundError:
            raise FileAccessError(path=path, mode="read")

        result = []
        for name in wb_xls.sheet_names():
            sheet = wb_xls.sheet_by_name(name)
            nrows = sheet.nrows
            # Count non-empty rows in first 5 rows
            nonempty = 0
            for row_idx in range(min(5, nrows)):
                row_cells = [sheet.cell(row_idx, c).value for c in range(sheet.ncols)]
                if any(c is not None and str(c).strip() != "" for c in row_cells):
                    nonempty += 1
            is_empty = nonempty < 2
            result.append(SheetInfo(name=name, approximate_row_count=nrows, is_empty=is_empty))
        wb_xls.release_resources()
        return result

    if ext == ".ods":
        try:
            with pd.ExcelFile(path, engine="odf") as xf:
                sheet_names: list[str] = list(xf.sheet_names)
        except PermissionError:
            raise FileAccessError(path=path, mode="read")
        except FileNotFoundError:
            raise FileAccessError(path=path, mode="read")

        result = []
        for name in sheet_names:
            try:
                df_head = pd.read_excel(path, engine="odf", sheet_name=name, nrows=5, header=None)
                nrows_approx = len(df_head)
                # Count rows with at least one non-null, non-empty cell
                nonempty = int(
                    df_head.apply(
                        lambda row: any(
                            v is not None and str(v).strip() != "" and str(v) != "nan"
                            for v in row
                        ),
                        axis=1,
                    ).sum()
                )
            except Exception:
                nrows_approx = 0
                nonempty = 0
            is_empty = nonempty < 2
            result.append(SheetInfo(name=name, approximate_row_count=nrows_approx, is_empty=is_empty))
        return result

    # For any other supported extension not handled above, raise UnsupportedFormatError
    raise UnsupportedFormatError(extension=ext)


# ---------------------------------------------------------------------------
# Public dispatch entry point
# ---------------------------------------------------------------------------


def read_input(
    path: pathlib.Path,
    sheet_name: str | None = None,
) -> ReadResult:
    """Dispatch entry point — picks the engine by file extension.

    Raises:
        UnsupportedFormatError: if the extension is not in SUPPORTED_EXTENSIONS.
        FileAccessError: if the file cannot be opened (PermissionError or
            FileNotFoundError from the underlying engine).

    Security note (T-1-02-05): extension whitelist check occurs before any
    file I/O. Unknown extensions never touch the file system.
    """
    ext = path.suffix.lower()
    if ext not in SUPPORTED_EXTENSIONS:
        raise UnsupportedFormatError(extension=ext or "(sem extensão)")

    if ext in {".xlsx", ".xlsm"}:
        return read_xlsx(path, sheet_name=sheet_name)

    if ext == ".xls":
        return read_xls(path, sheet_name=sheet_name)

    if ext == ".ods":
        return read_ods(path, sheet_name=sheet_name)

    if ext == ".csv":
        return read_csv_like(path, delimiter=",")

    if ext == ".tsv":
        return read_csv_like(path, delimiter="\t")

    # Unreachable: the whitelist check above covers all cases, but mypy needs this.
    raise UnsupportedFormatError(extension=ext)
