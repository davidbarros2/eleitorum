"""Synthetic fixture generators for the EleitorUM test suite.

All 15 functions listed below produce synthetic input files that downstream
plans consume. Each function embeds a specific quirk (encoding issue, whitespace
chaos, floating-point numeric IDs, etc.) so that the corresponding transformation
or validation rule can be exercised in isolation.

Per Eleitorum.md Section 14.1: all names include the words "Teste", "Exemplo",
or "Sintetica" so they are unmistakably synthetic and can never be confused with
real personal data.

Functions
---------
| # | Name                          | Quirk tested                              |
|---|-------------------------------|-------------------------------------------|
| 1 | make_simple_caderno           | Clean 2-column CSV (mec + name)           |
| 2 | make_simple_elegiveis         | Clean 2-column CSV (designation only)     |
| 3 | make_multi_sheet_xlsx         | 3-sheet workbook; empty Alunos sheet      |
| 4 | make_titled_xlsx              | Title row before real header              |
| 5 | make_headerless_xlsx          | No header row; data starts at row 1       |
| 6 | make_mojibake_csv             | UTF-8 bytes stored as Latin-1             |
| 7 | make_whitespace_chaos_xlsx    | Names with TAB, NBSP, ZWSP, spaces        |
| 8 | make_with_commas              | Names with trailing commas                |
| 9 | make_with_parentheses         | Names with parenthetical annotations      |
|10 | make_duplicate_within_prefix  | Same mec twice (VAL-03)                   |
|11 | make_cross_prefix_collision   | F500 + D500 together (VAL-04)             |
|12 | make_leading_zeros            | Mecs like F0500, D00123, b007             |
|13 | make_excel_float_numbers      | Numeric mec column stored as float        |
|14 | make_mixed_case_prefixes      | Mix of f6688 / F1234 (majority lowercase) |
|15 | make_unicode_replacement      | Name containing U+FFFD                    |

Per Eleitorum.md Section 14.3.
"""

import csv
import pathlib

import openpyxl

# ---------------------------------------------------------------------------
# Synthetic data constants
# (duplicated here intentionally so generators.py is importable without pytest)
# ---------------------------------------------------------------------------

SYNTHETIC_NAMES: tuple[str, ...] = (
    "João Silva Teste",
    "Maria Costa Exemplo",
    "Ana Pereira Sintetica",
    "Carlos Oliveira Teste",
    "Rui Ferreira Exemplo",
    "Sofia Santos Sintetica",
    "Marta Rodrigues Teste",
    "Pedro Martins Exemplo",
    "Inês Gomes Sintetica",
    "Tiago Lopes Teste",
    "Beatriz Cunha Exemplo",
    "Filipe Azevedo Sintetico",
)

# Valid mecanografico prefixes per D-08 (CONTEXT.md)
SYNTHETIC_PREFIXES: tuple[str, ...] = ("A", "PG", "ID", "F", "D", "B", "Q", "EX")

# Representative mec numbers for caderno fixtures
_CADERNO_ROWS: tuple[tuple[str, str], ...] = (
    ("f6688", "João Silva Teste"),
    ("f1234", "Maria Costa Exemplo"),
    ("f9001", "Ana Pereira Sintetica"),
    ("f2002", "Carlos Oliveira Teste"),
    ("d5500", "Rui Ferreira Exemplo"),
    ("d3311", "Sofia Santos Sintetica"),
    ("b7700", "Marta Rodrigues Teste"),
    ("b8800", "Pedro Martins Exemplo"),
    ("a100", "Inês Gomes Sintetica"),
    ("pg200", "Tiago Lopes Teste"),
    ("id300", "Beatriz Cunha Exemplo"),
    ("q400", "Filipe Azevedo Sintetico"),
    ("ex500", "João Silva Teste"),
    ("a101", "Maria Costa Exemplo"),
    ("pg201", "Ana Pereira Sintetica"),
    ("f6689", "Carlos Oliveira Teste"),
    ("f6690", "Rui Ferreira Exemplo"),
    ("d5501", "Sofia Santos Sintetica"),
    ("b7701", "Marta Rodrigues Teste"),
    ("ex501", "Pedro Martins Exemplo"),
)


# ---------------------------------------------------------------------------
# Helper
# ---------------------------------------------------------------------------


def _write_csv(
    path: pathlib.Path,
    rows: list[list[str]],
    encoding: str = "utf-8-sig",
) -> None:
    """Write rows to a UTF-8-with-BOM CSV using stdlib csv."""
    with open(path, "w", encoding=encoding, newline="") as fh:
        writer = csv.writer(
            fh,
            delimiter=";",
            quoting=csv.QUOTE_NONE,
            escapechar="\\",
            lineterminator="\r\n",
        )
        writer.writerows(rows)


# ---------------------------------------------------------------------------
# Fixture 1: simple caderno
# ---------------------------------------------------------------------------


def make_simple_caderno(path: pathlib.Path) -> pathlib.Path:
    """Write a clean 2-column caderno CSV (mec + name, UTF-8 BOM).

    Contains ~20 rows of synthetic data with mixed valid prefixes.
    Tests the happy-path reader and transformation pipeline.
    """
    rows: list[list[str]] = [["nº mec.", "nome"]]
    rows.extend([list(row) for row in _CADERNO_ROWS])
    _write_csv(path, rows)
    return path


# ---------------------------------------------------------------------------
# Fixture 2: simple elegíveis
# ---------------------------------------------------------------------------


def make_simple_elegiveis(path: pathlib.Path) -> pathlib.Path:
    """Write a clean elegíveis CSV (designation only column, UTF-8 BOM).

    No mec column — exercises the elegíveis pipeline path.
    """
    rows: list[list[str]] = [["nome"]]
    for name in SYNTHETIC_NAMES:
        rows.append([name])
    _write_csv(path, rows)
    return path


# ---------------------------------------------------------------------------
# Fixture 3: multi-sheet XLSX
# ---------------------------------------------------------------------------


def make_multi_sheet_xlsx(path: pathlib.Path) -> pathlib.Path:
    """Write a 3-sheet XLSX: Docentes (10 rows), PTAG (5 rows), Alunos (header only).

    The empty Alunos sheet exercises INP-11 (empty sheet detection).
    Sheet order is: Docentes, PTAG, Alunos.
    """
    wb = openpyxl.Workbook()
    # openpyxl creates a default "Sheet" — remove it
    default_ws = wb.active
    if default_ws is not None:
        wb.remove(default_ws)

    # Sheet 1: Docentes — 10 data rows
    ws_docentes = wb.create_sheet("Docentes")
    ws_docentes.append(["Nº Mec.", "Nome"])
    for i, (mec, name) in enumerate(_CADERNO_ROWS[:10]):
        ws_docentes.append([mec, name])

    # Sheet 2: PTAG — 5 data rows
    ws_ptag = wb.create_sheet("PTAG")
    ws_ptag.append(["Nº Mec.", "Nome"])
    for i, (mec, name) in enumerate(_CADERNO_ROWS[10:15]):
        ws_ptag.append([mec, name])

    # Sheet 3: Alunos — header only, no data rows (exercises INP-11)
    ws_alunos = wb.create_sheet("Alunos")
    ws_alunos.append(["Nº Mec.", "Nome"])

    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Fixture 4: titled XLSX (title row before real header)
# ---------------------------------------------------------------------------


def make_titled_xlsx(path: pathlib.Path) -> pathlib.Path:
    """Write an XLSX where row 0 is a title, headers are in row 2 (0-indexed).

    Tests DET-01 header-row scoring when the first row is not the header.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Folha1")
    ws.title = "Folha1"

    # Row 1: title
    ws.append(["Título: Cadernos Provisórios 2026", "", ""])
    # Row 2: empty separator
    ws.append(["", "", ""])
    # Row 3: actual header
    ws.append(["Nº Mec.", "Nome", ""])
    # Rows 4+: data
    for mec, name in _CADERNO_ROWS[:10]:
        ws.append([mec, name, ""])

    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Fixture 5: headerless XLSX
# ---------------------------------------------------------------------------


def make_headerless_xlsx(path: pathlib.Path) -> pathlib.Path:
    """Write an XLSX with NO header row — first row is data.

    Tests DET-02: when no header is found, pipeline signals manual mapping mode.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Folha1")
    ws.title = "Folha1"

    for mec, name in _CADERNO_ROWS[:10]:
        ws.append([mec, name])

    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Fixture 6: mojibake CSV
# ---------------------------------------------------------------------------


def make_mojibake_csv(path: pathlib.Path) -> pathlib.Path:
    """Write a CSV where UTF-8 strings were saved as Latin-1, producing mojibake.

    Construction: encode each name as UTF-8, re-interpret the raw bytes as
    Latin-1, then write those characters to a Latin-1 file. The on-disk bytes
    contain byte sequences like \\xc3\\xa3 (the UTF-8 encoding of 'ã') stored
    as two Latin-1 characters 'Ã£', which is the classic mojibake fingerprint.

    Tests TRF-09 (deterministic mojibake correction).
    """
    header_utf8 = "nº mec.;nome\r\n"
    rows_utf8 = [
        f"f6688;{name}\r\n"
        for name in [
            "João Silva Teste",
            "Maria Costa Exemplo",
            "Ana Pereira Sintetica",
        ]
    ]
    content_utf8 = header_utf8 + "".join(rows_utf8)

    # Simulate the mojibake: encode as UTF-8, then write those bytes as Latin-1
    mojibake_content = content_utf8.encode("utf-8").decode("latin-1")
    path.write_bytes(mojibake_content.encode("latin-1"))
    return path


# ---------------------------------------------------------------------------
# Fixture 7: whitespace chaos XLSX
# ---------------------------------------------------------------------------


def make_whitespace_chaos_xlsx(path: pathlib.Path) -> pathlib.Path:
    """Write an XLSX with whitespace-polluted names.

    Each name embeds: leading ASCII spaces, NO-BREAK SPACE (U+00A0),
    HORIZONTAL TAB (U+0009), ZERO-WIDTH SPACE (U+200B), and trailing spaces.

    Tests TRF-05 (strip all Unicode whitespace) and TRF-06 (collapse internal).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Folha1")
    ws.title = "Folha1"
    ws.append(["Nº Mec.", "Nome"])

    # Construct whitespace-polluted names using explicit code points
    chaos_names = [
        # two leading spaces + NBSP between first/last + TAB before suffix + ZWSP + two trailing
        "  Jo\xe3o Silva\tTeste​  ",
        # leading NBSP + internal ZWSP + trailing TAB
        " Maria​Costa Exemplo\t",
        # double spaces + NBSP throughout
        "Ana  Pereira Sintetica  ",
    ]
    mecs = ["f6688", "f1234", "f9001"]
    for mec, name in zip(mecs, chaos_names):
        ws.append([mec, name])

    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Fixture 8: names with commas
# ---------------------------------------------------------------------------


def make_with_commas(path: pathlib.Path) -> pathlib.Path:
    """Write a CSV where some names have trailing commas.

    Tests TRF-07 (comma removal from names).
    """
    rows: list[list[str]] = [
        ["nº mec.", "nome"],
        ["f6688", "Marta Oliveira Teste,"],
        ["f1234", "João Silva Teste"],
        ["f9001", "Maria Costa Exemplo,"],
        ["d5500", "Carlos Oliveira Teste"],
        ["b7700", "Ana Pereira Sintetica,"],
    ]
    _write_csv(path, rows)
    return path


# ---------------------------------------------------------------------------
# Fixture 9: names with parentheses
# ---------------------------------------------------------------------------


def make_with_parentheses(path: pathlib.Path) -> pathlib.Path:
    """Write a CSV where some names have parenthetical annotations.

    Tests TRF-08 (parenthesis removal and re-whitespacing).
    """
    rows: list[list[str]] = [
        ["nº mec.", "nome"],
        ["f6688", "Rui Pereira Teste (Coordenador)"],
        ["f1234", "João Silva Teste"],
        ["f9001", "Maria Costa Exemplo (Diretora de Curso)"],
        ["d5500", "Ana Pereira Sintetica (Docente)"],
        ["b7700", "Carlos Oliveira Teste"],
    ]
    _write_csv(path, rows)
    return path


# ---------------------------------------------------------------------------
# Fixture 10: duplicate within prefix
# ---------------------------------------------------------------------------


def make_duplicate_within_prefix(path: pathlib.Path) -> pathlib.Path:
    """Write a CSV with the same mec number appearing twice within prefix F.

    The mec f6688 appears twice with different names. This must trigger VAL-03
    (duplicate-within-prefix hard error).
    """
    rows: list[list[str]] = [
        ["nº mec.", "nome"],
        ["f6688", "João Silva Teste"],
        ["f1234", "Maria Costa Exemplo"],
        ["f6688", "Ana Pereira Sintetica"],  # duplicate of first row
        ["d5500", "Carlos Oliveira Teste"],
    ]
    _write_csv(path, rows)
    return path


# ---------------------------------------------------------------------------
# Fixture 11: cross-prefix collision (F/D/B share namespace)
# ---------------------------------------------------------------------------


def make_cross_prefix_collision(path: pathlib.Path) -> pathlib.Path:
    """Write a CSV with F500 and D500, which share the F/D/B numeric namespace.

    Per D-08, F, D, and B prefixes share a single uniqueness namespace, so
    F500 and D500 together trigger VAL-04 (cross-prefix collision hard error).
    """
    rows: list[list[str]] = [
        ["nº mec.", "nome"],
        ["F500", "João Silva Teste"],
        ["D500", "Maria Costa Exemplo"],
        ["f1234", "Ana Pereira Sintetica"],
        ["b9900", "Carlos Oliveira Teste"],
    ]
    _write_csv(path, rows)
    return path


# ---------------------------------------------------------------------------
# Fixture 12: leading zeros in mec
# ---------------------------------------------------------------------------


def make_leading_zeros(path: pathlib.Path) -> pathlib.Path:
    """Write a CSV with mecanográficos that have leading zeros.

    Tests TRF-03 (strip leading zeros from the numeric part of the mec).
    """
    rows: list[list[str]] = [
        ["nº mec.", "nome"],
        ["F0500", "João Silva Teste"],
        ["D00123", "Maria Costa Exemplo"],
        ["b007", "Ana Pereira Sintetica"],
        ["f006688", "Carlos Oliveira Teste"],
        ["pg0200", "Rui Ferreira Exemplo"],
    ]
    _write_csv(path, rows)
    return path


# ---------------------------------------------------------------------------
# Fixture 13: Excel float numeric mec column
# ---------------------------------------------------------------------------


def make_excel_float_numbers(path: pathlib.Path) -> pathlib.Path:
    """Write an XLSX where the mec column contains float cell values.

    Excel commonly stores integer IDs as floats (e.g., 14891 becomes 14891.0).
    Tests TRF-02 (float-to-int-string conversion).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Folha1")
    ws.title = "Folha1"
    ws.append(["Nº Mec.", "Nome"])

    float_mecs: list[tuple[float, str]] = [
        (14891.0, "João Silva Teste"),
        (6688.0, "Maria Costa Exemplo"),
        (1234.0, "Ana Pereira Sintetica"),
        (9001.0, "Carlos Oliveira Teste"),
        (5500.0, "Rui Ferreira Exemplo"),
    ]
    for mec_float, name in float_mecs:
        # Write the mec as a Python float so openpyxl persists it as a number
        ws.append([mec_float, name])

    wb.save(path)
    wb.close()
    return path


# ---------------------------------------------------------------------------
# Fixture 14: mixed-case prefixes
# ---------------------------------------------------------------------------


def make_mixed_case_prefixes(path: pathlib.Path) -> pathlib.Path:
    """Write a CSV with a mix of lowercase and uppercase F prefixes.

    The lowercase rows outnumber uppercase (3 lowercase vs 2 uppercase),
    so TRF-04 (majority-wins case normalization) should normalize to lowercase.
    """
    rows: list[list[str]] = [
        ["nº mec.", "nome"],
        ["f6688", "João Silva Teste"],       # lowercase
        ["F1234", "Maria Costa Exemplo"],    # uppercase
        ["f9001", "Ana Pereira Sintetica"],  # lowercase
        ["F2002", "Carlos Oliveira Teste"],  # uppercase
        ["f7003", "Rui Ferreira Exemplo"],   # lowercase
    ]
    _write_csv(path, rows)
    return path


# ---------------------------------------------------------------------------
# Fixture 15: Unicode replacement character in name
# ---------------------------------------------------------------------------


def make_unicode_replacement(path: pathlib.Path) -> pathlib.Path:
    """Write a CSV where a name contains the Unicode replacement character U+FFFD.

    Tests TRF-11 (remove U+FFFD from names, keep the rest of the name).
    """
    rows: list[list[str]] = [
        ["nº mec.", "nome"],
        ["f6688", "Jo�ão Silva Teste"],  # U+FFFD embedded in name
        ["f1234", "Maria Costa Exemplo"],
        ["f9001", "Ana Pere�ira Sintetica"],  # U+FFFD mid-word
    ]
    _write_csv(path, rows)
    return path
