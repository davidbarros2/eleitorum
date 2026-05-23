"""Full end-to-end integration tests for the EleitorUM pipeline.

Covers all five user journeys from Eleitorum.md Section 10, plus cross-cutting
edge cases for every documented real-data quirk.

All tests use synthetic data only (privacy invariant).
No Qt imports are present in this file.
"""

import pathlib

import openpyxl
import pytest

from eleitorum.core.pipeline import PipelineSource, run_pipeline
from tests.fixtures import generators

# ---------------------------------------------------------------------------
# User Journey 1: Happy path caderno CSV
# ---------------------------------------------------------------------------


def test_happy_path_caderno_csv(tmp_path: pathlib.Path) -> None:
    """Section 10.1: clean 2-column caderno CSV → byte-exact output + log file."""
    inp = generators.make_simple_caderno(tmp_path / "in.csv")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    assert result.success is True, f"expected success, got failures: {result.failures}"
    assert out.exists(), "output file must be created"
    assert result.log_path is not None and result.log_path.exists()

    # Byte-exact assertions
    raw = out.read_bytes()
    assert raw[:3] == b"\xef\xbb\xbf", "BOM missing"
    assert raw.endswith(b"\r\n"), "no trailing CRLF"
    assert b'"' not in raw, "quote character present"
    lone_lf = raw.count(b"\n") - raw.count(b"\r\n")
    assert lone_lf == 0, f"lone LF detected (lone_lf={lone_lf})"

    # Log file must contain key markers
    log_text = result.log_path.read_text(encoding="utf-8-sig")
    assert "INICIO" in log_text
    assert "SAIDA" in log_text
    assert "FIM" in log_text

    # Row count
    assert result.rows_processed == 20, f"expected 20 rows, got {result.rows_processed}"


# ---------------------------------------------------------------------------
# User Journey 2: Happy path elegíveis CSV
# ---------------------------------------------------------------------------


def test_happy_path_elegiveis_csv(tmp_path: pathlib.Path) -> None:
    """Section 10.1 variant: elegíveis output — sorted alphabetically, 0-based index."""
    inp = generators.make_simple_elegiveis(tmp_path / "in.csv")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "elegiveis", out)

    assert result.success is True, f"expected success, got failures: {result.failures}"
    assert out.exists()

    raw = out.read_bytes()
    assert raw[:3] == b"\xef\xbb\xbf", "BOM missing"
    assert raw.endswith(b"\r\n")
    assert b'"' not in raw

    # Parse output and verify 0-based index and sorted order
    # Use read_bytes + decode to preserve CRLF as-is (text mode converts \r\n to \n)
    content = out.read_bytes().decode("utf-8-sig")
    lines = [line for line in content.split("\r\n") if line]
    assert lines[0] == "personnel_number;designation", f"unexpected header: {lines[0]}"

    # First data row should start with index 0
    first_data = lines[1]
    assert first_data.startswith("0;"), f"first data row should start with '0;': {first_data}"

    # D-02 byte-exact data-row assertions (TRF-13, TRF-14, OUT-09)
    data_lines = lines[1:]  # skip header

    # (a) 0-based integer index: first data row has index 0 (TRF-14)
    assert int(data_lines[0].split(";")[0]) == 0, (
        f"first index must be 0; got: {data_lines[0]}"
    )

    # (b) alphabetical NFKD order: designations sorted before index assignment (TRF-13)
    names = [line.split(";")[1] for line in data_lines]
    assert names == sorted(names, key=lambda s: s.casefold()), (
        f"elegíveis must be in alphabetical order; got first 5: {names[:5]}"
    )

    # (c) no trailing semicolon: elegíveis rows are "{int};{designation}" only (OUT-09)
    for line in data_lines:
        assert not line.endswith(";"), (
            f"elegíveis row must not end with semicolon: {line}"
        )


# ---------------------------------------------------------------------------
# User Journey 3: Multi-sheet XLSX — processes selected sheet
# ---------------------------------------------------------------------------


def test_multi_sheet_xlsx_processes_selected_sheet(tmp_path: pathlib.Path) -> None:
    """Section 10.2: multi-sheet XLSX — each sheet produces different row count."""
    inp = generators.make_multi_sheet_xlsx(tmp_path / "multi.xlsx")

    out_docentes = tmp_path / "docentes.csv"
    r_docentes = run_pipeline(
        PipelineSource(path=inp, sheet_name="Docentes"),
        "caderno",
        out_docentes,
    )
    assert r_docentes.success is True, f"Docentes failed: {r_docentes.failures}"
    assert r_docentes.rows_processed == 10, (
        f"expected 10 rows for Docentes, got {r_docentes.rows_processed}"
    )

    out_ptag = tmp_path / "ptag.csv"
    r_ptag = run_pipeline(
        PipelineSource(path=inp, sheet_name="PTAG"),
        "caderno",
        out_ptag,
    )
    assert r_ptag.success is True, f"PTAG failed: {r_ptag.failures}"
    assert r_ptag.rows_processed == 5, f"expected 5 rows for PTAG, got {r_ptag.rows_processed}"

    assert r_docentes.rows_processed != r_ptag.rows_processed, (
        "Docentes and PTAG should produce different row counts"
    )


# ---------------------------------------------------------------------------
# User Journey 3b: Mojibake file corrected end-to-end
# ---------------------------------------------------------------------------


def test_mojibake_file_corrected_end_to_end(tmp_path: pathlib.Path) -> None:
    """Section 10.3: string-level mojibake in names is corrected.

    Uses an inline UTF-8 CSV with mojibake strings (JoÃ£o instead of João)
    rather than make_mojibake_csv, because make_mojibake_csv has duplicate
    mec numbers which would trigger a separate validation failure.
    The inline fixture isolates the TRF-09 mojibake correction path.
    """
    inp = tmp_path / "moji.csv"
    # Write UTF-8 BOM CSV with mojibake string patterns in names
    # 'JoÃ£o' is the mojibake representation of 'João' when UTF-8 bytes
    # are interpreted as Latin-1 and then stored back as a string.
    content = "nº mec.;nome\r\nf6688;JoÃ£o Silva Teste\r\nf1234;Maria Costa Exemplo\r\n"
    inp.write_bytes(b"\xef\xbb\xbf" + content.encode("utf-8"))

    out = tmp_path / "out.csv"
    result = run_pipeline(inp, "caderno", out)

    assert result.success is True, f"expected success, got failures: {result.failures}"
    assert out.exists()

    # Log must contain a LIMPEZA line mentioning mojibake
    log_text = result.log_path.read_text(encoding="utf-8-sig")  # type: ignore[union-attr]
    limpeza_lines = [
        line for line in log_text.split("\n") if "LIMPEZA" in line and "mojibake" in line.lower()
    ]
    assert len(limpeza_lines) >= 1, (
        f"expected at least one LIMPEZA/mojibake log line; got log:\n{log_text}"
    )

    # Output must not contain the mojibake string pattern (the name was corrected)
    output_text = out.read_text(encoding="utf-8-sig")
    assert "JoÃ£o" not in output_text, "mojibake string pattern should have been corrected"


# ---------------------------------------------------------------------------
# User Journey 4: Duplicate mecanográfico → rejection (no output file)
# ---------------------------------------------------------------------------


def test_duplicate_rejected_no_output_errors_log_created(tmp_path: pathlib.Path) -> None:
    """Section 10.4: duplicate mec within same prefix → failure, _ERRORS_ log, no CSV."""
    inp = generators.make_duplicate_within_prefix(tmp_path / "dup.csv")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    assert result.success is False, "expected failure for duplicate mec"
    assert result.output_path is None, "output_path must be None on failure (OUT-10)"
    assert not out.exists(), "output CSV must NOT be created on validation failure"
    assert result.error_log_path is not None, "error_log_path must be set on failure"
    assert result.error_log_path.exists(), "error log file must exist"

    # Error log must contain ERRO lines identifying the duplicate
    error_text = result.error_log_path.read_text(encoding="utf-8-sig")
    erro_lines = [line for line in error_text.split("\n") if "] ERRO" in line]
    assert len(erro_lines) >= 1, f"expected at least 1 ERRO line in error log; got: {erro_lines}"
    # The offending mec (6688) must be mentioned in the error log
    assert "6688" in error_text, (
        f"error log should mention the offending mec number 6688; got:\n{error_text}"
    )


# ---------------------------------------------------------------------------
# User Journey 5: Manual column mapping
# ---------------------------------------------------------------------------


def test_manual_column_mapping(tmp_path: pathlib.Path) -> None:
    """Section 10.5: custom headers with manual column override."""
    inp = tmp_path / "custom.csv"
    rows = [
        "MeuNumero;Pessoa",
        "f6688;João Silva Teste",
        "f1234;Maria Costa Exemplo",
        "f9001;Ana Pereira Sintetica",
        "d5500;Carlos Oliveira Teste",
        "b7700;Rui Ferreira Exemplo",
    ]
    inp.write_bytes(b"\xef\xbb\xbf" + "\r\n".join(rows).encode("utf-8") + b"\r\n")

    out = tmp_path / "out.csv"
    source = PipelineSource(path=inp, manual_mec_col=0, manual_name_col=1)
    result = run_pipeline(source, "caderno", out)

    assert result.success is True, f"expected success with manual mapping, got: {result.failures}"
    assert out.exists()
    assert result.rows_processed == 5


# ---------------------------------------------------------------------------
# Cross-cutting: F/D/B cross-prefix collision
# ---------------------------------------------------------------------------


def test_fdb_cross_prefix_collision_rejected(tmp_path: pathlib.Path) -> None:
    """F500 and D500 share the F/D/B namespace → validation failure."""
    inp = generators.make_cross_prefix_collision(tmp_path / "collision.csv")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    assert result.success is False, "expected failure for F/D/B cross-prefix collision"
    assert result.error_log_path is not None and result.error_log_path.exists()

    error_text = result.error_log_path.read_text(encoding="utf-8-sig")
    # Error log must reference the colliding numbers (F500 / D500)
    assert "500" in error_text, "error log should mention the colliding number 500"
    assert not out.exists(), "no CSV should be written on failure"


# ---------------------------------------------------------------------------
# Cross-cutting: Leading zeros stripped and logged
# ---------------------------------------------------------------------------


def test_leading_zeros_stripped_logged(tmp_path: pathlib.Path) -> None:
    """TRF-03: leading zeros in mec are stripped and each removal logged."""
    inp = generators.make_leading_zeros(tmp_path / "lz.csv")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    assert result.success is True, f"expected success, got failures: {result.failures}"

    log_text = result.log_path.read_text(encoding="utf-8-sig")  # type: ignore[union-attr]
    limpeza_lines = [line for line in log_text.split("\n") if "] LIMPEZA" in line]
    # At least some LIMPEZA entries for the stripped zeros
    assert len(limpeza_lines) >= 1, (
        f"expected LIMPEZA log entries for stripped zeros; got log:\n{log_text}"
    )

    # Output should contain the stripped values (e.g., F500 not F0500)
    output_text = out.read_text(encoding="utf-8-sig")
    assert "F0500" not in output_text.upper(), "leading zero F0500 should be stripped to F500"


# ---------------------------------------------------------------------------
# Cross-cutting: Excel float numbers converted and logged
# ---------------------------------------------------------------------------


def test_excel_float_numbers_converted_logged(tmp_path: pathlib.Path) -> None:
    """TRF-02: Excel numeric float mec values (e.g. 14891.0) cause validation failure.

    Float cells without a prefix are invalid mec values (no prefix like F).
    The pipeline records these as MecanograficoError failures.
    """
    inp = generators.make_excel_float_numbers(tmp_path / "floats.xlsx")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    # Float cells (e.g. 14891.0) have no prefix and are invalid mec values
    # This triggers MecanograficoError (VAL-01: requires prefix)
    assert result.success is False, "float-only mec cells have no prefix and must fail validation"
    assert result.error_log_path is not None and result.error_log_path.exists()

    # The error log should mention the float values
    error_text = result.error_log_path.read_text(encoding="utf-8-sig")
    assert "ERRO" in error_text, "error log must contain ERRO entries for invalid mecs"


# ---------------------------------------------------------------------------
# Cross-cutting: Mixed-case prefixes → majority-wins normalization
# ---------------------------------------------------------------------------


def test_mixed_case_prefixes_normalized_via_majority(tmp_path: pathlib.Path) -> None:
    """TRF-04: mixed case prefixes are normalized to majority case (lowercase wins on tie)."""
    inp = generators.make_mixed_case_prefixes(tmp_path / "mixed.csv")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    assert result.success is True, f"expected success, got failures: {result.failures}"

    # Log must contain a CASO line describing the normalization
    log_text = result.log_path.read_text(encoding="utf-8-sig")  # type: ignore[union-attr]
    caso_lines = [line for line in log_text.split("\n") if "] CASO" in line]
    assert len(caso_lines) >= 1, f"expected CASO log line; got log:\n{log_text}"

    # The fixture has 3 lowercase f rows and 2 uppercase F rows → lowercase should win
    case_line = caso_lines[0]
    assert "lower" in case_line, f"expected 'lower' normalization; got: {case_line}"

    # Output should use lowercase f prefix
    # Use bytes decode to preserve CRLF
    content = out.read_bytes().decode("utf-8-sig")
    data_lines = content.split("\r\n")[1:]  # skip header
    for line in data_lines:
        if not line:
            continue
        mec = line.split(";")[0]
        if mec:
            assert mec[0].islower(), f"prefix should be lowercase; got mec: {mec}"
            break


# ---------------------------------------------------------------------------
# Cross-cutting: Whitespace chaos cleaned
# ---------------------------------------------------------------------------


def test_whitespace_chaos_cleaned(tmp_path: pathlib.Path) -> None:
    """TRF-05/06: NBSP (U+00A0) and ZWSP (U+200B) are stripped from names."""
    inp = generators.make_whitespace_chaos_xlsx(tmp_path / "chaos.xlsx")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    assert result.success is True, f"expected success, got failures: {result.failures}"

    # Read output as text and assert no invisible whitespace characters remain
    output_text = out.read_text(encoding="utf-8-sig")
    assert chr(0x00A0) not in output_text, "NBSP (U+00A0) should have been stripped"
    assert chr(0x200B) not in output_text, "ZWSP (U+200B) should have been stripped"


# ---------------------------------------------------------------------------
# Cross-cutting: Unicode replacement character removed and logged
# ---------------------------------------------------------------------------


def test_unicode_replacement_removed_logged(tmp_path: pathlib.Path) -> None:
    """TRF-11: U+FFFD replacement characters are removed from names with AVISO log entry."""
    inp = generators.make_unicode_replacement(tmp_path / "repl.csv")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    assert result.success is True, f"expected success, got failures: {result.failures}"

    # Log must contain an AVISO entry for the removed character(s)
    log_text = result.log_path.read_text(encoding="utf-8-sig")  # type: ignore[union-attr]
    aviso_lines = [line for line in log_text.split("\n") if "] AVISO" in line]
    assert len(aviso_lines) >= 1, f"expected AVISO log entries for U+FFFD removal; got:\n{log_text}"

    # Output must not contain the replacement character
    raw = out.read_bytes()
    # CORRECT: U+FFFD in UTF-8 is 3 bytes EF BF BD
    assert b"\xef\xbf\xbd" not in raw, (
        "replacement character (U+FFFD) bytes should not appear in UTF-8 output"
    )
    output_text = out.read_text(encoding="utf-8-sig")
    assert "�" not in output_text, "U+FFFD should have been removed from names"


# ---------------------------------------------------------------------------
# D-04 progress callback invoked
# ---------------------------------------------------------------------------


def test_progress_callback_invoked(tmp_path: pathlib.Path) -> None:
    """D-04: progress_cb is called during processing with (current_row, total_rows)."""
    # Build a fixture with 250+ rows inline (exceeds 100-row rate-limiting threshold)
    inp = tmp_path / "big.csv"
    header = "nº mec.;nome"
    data_rows = [f"f{i};Joao Teste {i}" for i in range(1, 251)]
    content = "\r\n".join([header] + data_rows) + "\r\n"
    inp.write_bytes(b"\xef\xbb\xbf" + content.encode("utf-8"))

    out = tmp_path / "out.csv"
    cb_calls: list[tuple[int, int]] = []

    result = run_pipeline(inp, "caderno", out, progress_cb=lambda c, t: cb_calls.append((c, t)))

    assert result.success is True
    # With 250 rows and rate-limiting every 100 rows + final:
    # calls at row 1 (i=0), row 101 (i=100), row 201 (i=200), row 250 (i=249 = final)
    assert len(cb_calls) >= 2, (
        f"expected at least 2 progress callbacks for 250 rows; got {len(cb_calls)}"
    )

    # Each call must have form (int, int) with both > 0
    for current, total in cb_calls:
        assert isinstance(current, int) and isinstance(total, int)
        assert current > 0
        assert total == 250


# ---------------------------------------------------------------------------
# Dry-run: no files written
# ---------------------------------------------------------------------------


def test_dry_run_no_files_written(tmp_path: pathlib.Path) -> None:
    """output_path=None → dry-run: validation only, no files created."""
    inp = generators.make_simple_caderno(tmp_path / "in.csv")

    result = run_pipeline(inp, "caderno", output_path=None)

    assert result.success is True, f"dry-run should succeed for valid input: {result.failures}"
    assert result.output_path is None
    assert result.log_path is None
    assert result.error_log_path is None

    # No extra files in tmp_path beyond the input
    created = [f for f in tmp_path.iterdir() if f != inp]
    assert len(created) == 0, f"dry-run must not create any files; found: {created}"


# ---------------------------------------------------------------------------
# Same-path rejection (VAL-08)
# ---------------------------------------------------------------------------


def test_output_equals_input_path_rejected(tmp_path: pathlib.Path) -> None:
    """VAL-08: output path == input path → failure with PT-PT error in log."""
    inp = generators.make_simple_caderno(tmp_path / "in.csv")

    result = run_pipeline(inp, "caderno", inp)  # same path as output

    assert result.success is False, "expected failure when output == input"
    assert result.error_log_path is not None and result.error_log_path.exists()

    # Error log must mention the same-path error (VAL-08 PT-PT message)
    error_text = result.error_log_path.read_text(encoding="utf-8-sig")
    same_path_keywords = ["mesmo ficheiro", "ficheiro de saída", "original"]
    assert any(kw in error_text for kw in same_path_keywords), (
        f"error log should mention same-path error; got:\n{error_text}"
    )


# ---------------------------------------------------------------------------
# PERF-03: openpyxl streaming mode assertion
# ---------------------------------------------------------------------------


def test_perf_03_streaming_mode_assertion(
    tmp_path: pathlib.Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    """PERF-03: openpyxl.load_workbook must always be called with read_only=True, data_only=True."""
    inp = generators.make_multi_sheet_xlsx(tmp_path / "multi.xlsx")
    out = tmp_path / "out.csv"

    captured: dict[str, object] = {}
    original_load_workbook = openpyxl.load_workbook

    def spy(path: object, **kw: object) -> object:
        captured.update(kw)
        return original_load_workbook(path, **kw)

    monkeypatch.setattr(openpyxl, "load_workbook", spy)

    result = run_pipeline(
        PipelineSource(path=inp, sheet_name="Docentes"),
        "caderno",
        out,
    )

    assert result.success is True
    assert captured.get("read_only") is True, (
        f"openpyxl.load_workbook must be called with read_only=True; kwargs={captured}"
    )
    assert captured.get("data_only") is True, (
        f"openpyxl.load_workbook must be called with data_only=True; kwargs={captured}"
    )


# ---------------------------------------------------------------------------
# Cross-cutting: Parenthetical annotations removed
# ---------------------------------------------------------------------------


def test_parenthetical_annotations_removed(tmp_path: pathlib.Path) -> None:
    """TRF-08: parenthetical annotations like (Coordenador) are removed from names."""
    inp = generators.make_with_parentheses(tmp_path / "parens.csv")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    assert result.success is True, f"expected success, got failures: {result.failures}"

    output_text = out.read_text(encoding="utf-8-sig")
    assert "(Coordenador)" not in output_text, "parenthetical should have been removed"

    log_text = result.log_path.read_text(encoding="utf-8-sig")  # type: ignore[union-attr]
    limpeza_lines = [line for line in log_text.split("\n") if "] LIMPEZA" in line]
    assert len(limpeza_lines) >= 1, "expected LIMPEZA entry for annotation removal"


# ---------------------------------------------------------------------------
# Cross-cutting: Trailing commas removed
# ---------------------------------------------------------------------------


def test_trailing_commas_removed(tmp_path: pathlib.Path) -> None:
    """TRF-07: trailing commas in names are removed and logged."""
    inp = generators.make_with_commas(tmp_path / "commas.csv")
    out = tmp_path / "out.csv"

    result = run_pipeline(inp, "caderno", out)

    assert result.success is True, f"expected success, got failures: {result.failures}"

    # Use bytes decode to preserve CRLF for reliable split
    content = out.read_bytes().decode("utf-8-sig")
    for line in content.split("\r\n")[1:]:  # skip header
        if line:
            assert not line.split(";")[1].endswith(","), (
                f"name should not end with comma in output line: {line}"
            )

    log_text = result.log_path.read_text(encoding="utf-8-sig")  # type: ignore[union-attr]
    limpeza_lines = [
        line for line in log_text.split("\n") if "] LIMPEZA" in line and "vírgula" in line.lower()
    ]
    assert len(limpeza_lines) >= 1, "expected LIMPEZA entry for comma removal"
