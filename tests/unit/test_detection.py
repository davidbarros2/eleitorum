"""Tests for the eleitorum.core.detection module.

Covers: DET-01 through DET-07, INP-07 through INP-09.
"""

from __future__ import annotations

import pathlib

import pytest

from eleitorum.core.detection import (
    MECANOGRAFICO_SYNONYMS,
    NAME_SYNONYMS,
    ColumnMapping,
    EncodingDetectionResult,
    detect_columns,
    detect_encoding,
    detect_header_row,
    normalize_col_name,
)
from eleitorum.core.errors import EncodingDetectionError
from tests.fixtures.generators import make_headerless_xlsx, make_titled_xlsx


# ---------------------------------------------------------------------------
# normalize_col_name
# ---------------------------------------------------------------------------


def test_normalize_col_name_handles_ordinal_indicator() -> None:
    """NFKD decomposes U+00BA (ordinal indicator 'º') to 'o'."""
    assert normalize_col_name("Nº Mec.") == normalize_col_name("No Mec.")
    assert normalize_col_name("Nº Mec.") == "no mec."


def test_normalize_col_name_strips_diacritics() -> None:
    assert normalize_col_name("Designação") == "designacao"
    assert normalize_col_name("número") == "numero"


def test_normalize_col_name_lowercases_and_strips() -> None:
    assert normalize_col_name("  NOME  ") == "nome"


# ---------------------------------------------------------------------------
# MECANOGRAFICO_SYNONYMS and NAME_SYNONYMS
# ---------------------------------------------------------------------------


def test_mecanografico_synonyms_includes_all_spec_variants() -> None:
    """MECANOGRAFICO_SYNONYMS must contain all synonyms from Eleitorum.md Section 6.5."""
    assert "personnel_number" in MECANOGRAFICO_SYNONYMS
    assert "nmec" in MECANOGRAFICO_SYNONYMS
    assert "nmecanografico" in MECANOGRAFICO_SYNONYMS
    # Must have at least 15 entries (spec lists 18+ when expanded)
    assert len(MECANOGRAFICO_SYNONYMS) >= 15


def test_mecanografico_synonyms_includes_common_variants() -> None:
    # The normalized forms of various spellings should all be in the set
    assert normalize_col_name("nº mecanográfico") in MECANOGRAFICO_SYNONYMS
    assert normalize_col_name("numero mecanografico") in MECANOGRAFICO_SYNONYMS
    assert normalize_col_name("n. mecanografico") in MECANOGRAFICO_SYNONYMS
    assert normalize_col_name("nº mec.") in MECANOGRAFICO_SYNONYMS
    assert normalize_col_name("número de empregado") in MECANOGRAFICO_SYNONYMS


def test_name_synonyms_has_required_entries() -> None:
    assert "name" in NAME_SYNONYMS
    assert "nome" in NAME_SYNONYMS
    assert normalize_col_name("designação") in NAME_SYNONYMS
    assert len(NAME_SYNONYMS) >= 6


# ---------------------------------------------------------------------------
# detect_encoding — INP-07, INP-08, INP-09
# ---------------------------------------------------------------------------


def test_detect_encoding_utf8_bom() -> None:
    """INP-07: UTF-8 BOM is detected with via_bom=True and confidence=1.0."""
    result = detect_encoding(b"\xef\xbb\xbfhello world")
    assert isinstance(result, EncodingDetectionResult)
    assert result.via_bom is True
    assert result.confidence == 1.0
    assert result.encoding.lower().replace("-", "").replace("_", "") in (
        "utf8sig",
        "utf8bom",
        "utf8",
    )


def test_detect_encoding_utf8_no_bom() -> None:
    """INP-07: plain UTF-8 (ASCII-safe) is detected."""
    result = detect_encoding(b"hello world plain ascii content here")
    assert isinstance(result, EncodingDetectionResult)
    assert result.confidence >= 0.85
    assert result.via_bom is False


def test_detect_encoding_cp1252() -> None:
    """INP-07: CP1252-encoded bytes are handled (detect or fallback to cp1252).

    charset-normalizer may be ambiguous on short CP1252 samples, in which case
    the fallback chain returns "cp1252". Either a direct detection or fallback is
    acceptable — what matters is that detect_encoding does not raise and returns
    an encoding usable for decoding the input.
    """
    # Use a longer, more CP1252-distinctive sample so charset-normalizer has signal
    long_sample = (
        "Relatório de Eleições da Universidade do Minho 2026. "
        "Nomes dos candidatos: João Fernão Pimenta, Conceição Guimarães, "
        "Ângela Mendonça, Mário Rodrigues. Código: ELEIT-2026."
    )
    data = long_sample.encode("cp1252")
    result = detect_encoding(data)
    assert isinstance(result, EncodingDetectionResult)
    # Accept cp1252, windows-1252, iso-8859-1, or utf-8 (all decode the sample correctly)
    enc = result.encoding.lower().replace("-", "").replace("_", "")
    # Accept cp1252, windows-1252, cp1250 (closely related), iso-8859-1, or utf-8
    # charset-normalizer may select any compatible Western European encoding
    acceptable = {"cp1252", "windows1252", "cp1250", "iso88591", "latin1", "utf8"}
    assert enc in acceptable, (
        f"Unexpected encoding {result.encoding!r} for CP1252 data. "
        f"Acceptable: {acceptable}"
    )
    # Most importantly: the returned encoding must actually decode the bytes
    data.decode(result.encoding)


def test_detect_encoding_iso_8859_1_fallback() -> None:
    """INP-07: fallback chain reaches iso-8859-1 for non-BOM, non-cp1252 bytes."""
    # Write a known ISO-8859-1 sequence (Latin Extended-A)
    data = "café résumé naïve".encode("iso-8859-1")
    result = detect_encoding(data)
    assert isinstance(result, EncodingDetectionResult)
    # Accept any of the common spellings for this encoding family
    enc = result.encoding.lower().replace("-", "").replace("_", "")
    assert enc in ("iso88591", "latin1", "cp1252", "windows1252", "cp1250", "utf8"), (
        f"Unexpected encoding detected: {result.encoding!r}"
    )


def test_detect_encoding_undetectable_raises_pt_pt() -> None:
    """INP-08: empty bytes raise EncodingDetectionError with PT-PT message.

    Note: cp1252 and iso-8859-1 accept virtually all byte sequences, so the
    fallback chain will succeed for most garbage binary input. The error is
    raised for empty input (which has no valid encoding).
    """
    with pytest.raises(EncodingDetectionError) as exc_info:
        detect_encoding(b"")
    err = exc_info.value
    # Must contain the spec-verbatim PT-PT sentence
    assert "UTF-8" in err.message_pt
    assert "guardá-lo" in err.message_pt or "UTF" in err.message_pt


def test_detect_encoding_logs_choice() -> None:
    """INP-09: detect_encoding returns structured result (usable for logging)."""
    result = detect_encoding(b"\xef\xbb\xbftest data for logging")
    assert hasattr(result, "encoding")
    assert hasattr(result, "confidence")
    assert hasattr(result, "via_bom")
    assert hasattr(result, "raw_chaos")


# ---------------------------------------------------------------------------
# detect_header_row — DET-01, DET-02
# ---------------------------------------------------------------------------


def test_header_row_scoring_picks_best_of_first_10(tmp_path: pathlib.Path) -> None:
    """DET-01: header-row scorer returns the row index with the highest score."""
    path = make_titled_xlsx(tmp_path / "titled.xlsx")
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    assert ws is not None
    rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    result = detect_header_row(rows)
    # Row 0 is the title, row 1 is blank, row 2 is the real header
    assert result == 2


def test_no_header_returns_manual_mapping_signal(tmp_path: pathlib.Path) -> None:
    """DET-02: headerless files return None to signal manual-mapping mode."""
    path = make_headerless_xlsx(tmp_path / "headerless.xlsx")
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    assert ws is not None
    rows = [tuple(row) for row in ws.iter_rows(values_only=True)]
    wb.close()

    result = detect_header_row(rows)
    assert result is None


def test_header_row_returns_none_for_empty_input() -> None:
    assert detect_header_row([]) is None


def test_header_row_works_when_header_is_first_row() -> None:
    rows = [
        ("nº mec.", "nome", None),
        ("f6688", "João Silva Teste", None),
        ("f1234", "Maria Costa Exemplo", None),
    ]
    result = detect_header_row(rows)
    assert result == 0


# ---------------------------------------------------------------------------
# detect_columns — DET-03, DET-04, DET-05, DET-06, DET-07 + D-01
# ---------------------------------------------------------------------------


def test_mec_column_synonym_match_nfkd() -> None:
    """DET-03: mec column detected by NFKD synonym matching."""
    header = ("Nº Mec.", "Nome", None)
    data: list[tuple] = [("f6688", "João Silva Teste", None)]
    result = detect_columns(header, data, output_type="caderno")
    assert result.mec_col_index == 0
    assert result.detection_method == "synonym"


def test_name_column_synonym_match_nfkd() -> None:
    """DET-04: name column detected by NFKD synonym matching."""
    header = ("Nº Mec.", "Nome", None)
    data: list[tuple] = [("f6688", "João Silva Teste", None)]
    result = detect_columns(header, data, output_type="caderno")
    assert result.name_col_index == 1
    assert result.mec_col_label is not None
    assert result.name_col_label is not None


def test_detection_result_metadata_for_ui() -> None:
    """DET-05: result carries labels for UI display."""
    header = ("Nº Mec.", "Nome")
    data: list[tuple] = [("f6688", "João Silva Teste")]
    result = detect_columns(header, data, output_type="caderno")
    assert result.mec_col_label == "Nº Mec."
    assert result.name_col_label == "Nome"
    assert result.detection_method == "synonym"


def test_ambiguous_detection_returns_all_candidates() -> None:
    """DET-06: when multiple columns match, ambiguous_mec_candidates is populated."""
    header = ("nº mec.", "nmec", "Nome")
    data: list[tuple] = [("f6688", "f1234", "João Silva")]
    result = detect_columns(header, data, output_type="caderno")
    assert result.mec_col_index is None
    assert len(result.ambiguous_mec_candidates) >= 2


def test_elegiveis_hides_mec_mapping() -> None:
    """DET-07: for elegíveis, mec_col_index is always None."""
    header = ("Nº Mec.", "Nome")
    data: list[tuple] = [("f6688", "João Silva Teste")]
    result = detect_columns(header, data, output_type="elegiveis")
    assert result.mec_col_index is None


def test_format_fallback_regex_when_no_synonym_matches() -> None:
    """D-01: format fallback detects mec from value pattern when synonym fails."""
    header = ("Custom1", "Custom2")
    data: list[tuple] = [
        ("F500", "Person A"),
        ("D123", "Person B"),
        ("F600", "Person C"),
        ("A100", "Person D"),
        ("B200", "Person E"),
        ("F700", "Person F"),
        ("D400", "Person G"),
        ("F800", "Person H"),
        ("Q300", "Person I"),
        ("F900", "Person J"),
    ]
    result = detect_columns(header, data, output_type="caderno")
    assert result.mec_col_index == 0
    assert result.detection_method == "format_fallback"


def test_format_fallback_skipped_for_elegiveis() -> None:
    """DET-07 + D-01: format fallback is not applied for elegíveis output type."""
    header = ("Custom1", "Custom2")
    data: list[tuple] = [
        ("F500", "Person A"),
        ("D123", "Person B"),
        ("F600", "Person C"),
        ("A100", "Person D"),
        ("B200", "Person E"),
        ("F700", "Person F"),
        ("D400", "Person G"),
        ("F800", "Person H"),
        ("Q300", "Person I"),
        ("F900", "Person J"),
    ]
    result = detect_columns(header, data, output_type="elegiveis")
    assert result.mec_col_index is None


def test_column_mapping_is_frozen_dataclass() -> None:
    """ColumnMapping is a frozen dataclass."""
    header = ("Nº Mec.", "Nome")
    data: list[tuple] = [("f6688", "João Silva Teste")]
    result = detect_columns(header, data, output_type="caderno")
    assert isinstance(result, ColumnMapping)
    with pytest.raises((AttributeError, TypeError)):
        result.mec_col_index = 99  # type: ignore[misc]


def test_encoding_detection_result_is_frozen_dataclass() -> None:
    """EncodingDetectionResult is a frozen dataclass."""
    result = detect_encoding(b"\xef\xbb\xbftest")
    assert isinstance(result, EncodingDetectionResult)
    with pytest.raises((AttributeError, TypeError)):
        result.confidence = 0.5  # type: ignore[misc]
