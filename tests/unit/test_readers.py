"""Tests for the eleitorum.core.readers module.

Covers: INP-01 through INP-13 (except INP-07/INP-08/INP-09 which belong to
detection.py in plan 03).
"""

from __future__ import annotations

import pathlib
from typing import Any
from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd
import pytest

from eleitorum.core.errors import FileAccessError, UnsupportedFormatError
from eleitorum.core.readers import (
    SUPPORTED_EXTENSIONS,
    ReadResult,
    SheetInfo,
    _strip_trailing_empty,
    list_sheets,
    read_csv_like,
    read_input,
    read_ods,
    read_xlsx,
)
from tests.fixtures.generators import make_multi_sheet_xlsx, make_simple_caderno

# ---------------------------------------------------------------------------
# SUPPORTED_EXTENSIONS sanity check
# ---------------------------------------------------------------------------


def test_supported_extensions_contains_all_six() -> None:
    assert frozenset({".xlsx", ".xlsm", ".xls", ".ods", ".csv", ".tsv"}) == SUPPORTED_EXTENSIONS


# ---------------------------------------------------------------------------
# XLSX reading (INP-01, PERF-03)
# ---------------------------------------------------------------------------


def test_read_xlsx_basic(tmp_path: pathlib.Path) -> None:
    """INP-01: read_xlsx against make_multi_sheet_xlsx Docentes sheet yields >= 10 rows."""
    xlsx_path = make_multi_sheet_xlsx(tmp_path / "multi.xlsx")
    result = read_xlsx(xlsx_path, sheet_name="Docentes")
    assert isinstance(result, ReadResult)
    assert isinstance(result.rows, list)
    assert len(result.rows) >= 10  # 10 data rows + 1 header
    # Each row must be a tuple
    assert all(isinstance(r, tuple) for r in result.rows)
    assert result.sheet_name == "Docentes"
    assert result.raw_bytes_sample is None


def test_read_xlsx_default_first_sheet(tmp_path: pathlib.Path) -> None:
    """INP-01: read_xlsx picks first sheet when sheet_name is None."""
    xlsx_path = make_multi_sheet_xlsx(tmp_path / "multi.xlsx")
    result = read_xlsx(xlsx_path)
    assert result.sheet_name == "Docentes"


def test_xlsx_uses_read_only_and_data_only(tmp_path: pathlib.Path) -> None:
    """PERF-03: read_xlsx must use openpyxl.load_workbook(read_only=True, data_only=True)."""
    xlsx_path = make_multi_sheet_xlsx(tmp_path / "multi.xlsx")

    captured_kwargs: dict[str, Any] = {}
    original_load = openpyxl.load_workbook

    def capturing_load_workbook(*args: Any, **kwargs: Any) -> Any:
        captured_kwargs.update(kwargs)
        return original_load(*args, **kwargs)

    with patch(
        "eleitorum.core.readers.openpyxl.load_workbook",
        side_effect=capturing_load_workbook,
    ):
        read_xlsx(xlsx_path)

    assert captured_kwargs.get("read_only") is True, "openpyxl must use read_only=True (PERF-03)"
    assert captured_kwargs.get("data_only") is True, "openpyxl must use data_only=True (PERF-03)"


def test_permission_error_on_locked_file(tmp_path: pathlib.Path) -> None:
    """INP-13: PermissionError from openpyxl is re-raised as FileAccessError(mode='read')."""
    xlsx_path = tmp_path / "locked.xlsx"
    xlsx_path.write_bytes(b"")  # file must exist for the path to be valid

    with (
        patch(
            "eleitorum.core.readers.openpyxl.load_workbook",
            side_effect=PermissionError("locked"),
        ),
        pytest.raises(FileAccessError) as exc_info,
    ):
        read_xlsx(xlsx_path)

    err = exc_info.value
    assert "aberto" in err.message_pt or "abrir" in err.message_pt
    # Must not contain "gravar" (that is the write mode message)
    assert "gravar" not in err.message_pt


def test_read_xlsx_file_not_found_raises_file_access_error(tmp_path: pathlib.Path) -> None:
    """INP-13: FileNotFoundError is also wrapped as FileAccessError."""
    missing_path = tmp_path / "nonexistent.xlsx"
    with pytest.raises(FileAccessError):
        read_xlsx(missing_path)


# ---------------------------------------------------------------------------
# Legacy XLS reading (INP-02)
# ---------------------------------------------------------------------------


def test_read_xls_legacy(tmp_path: pathlib.Path) -> None:
    # NOTE: xlrd 2.x supports XLS binary only and xlwt (XLS writer) is
    # deprecated/unmaintained and not in the project dependencies. Without a
    # legacy .xls writer, generating a synthetic .xls file in the test
    # environment is not possible via pure Python. This test is skipped with
    # explicit documentation; it will be covered in Phase 3 when a sample .xls
    # file is added to tests/fixtures/sample_files/.
    pytest.skip(
        "Legacy XLS fixture requires xlwt which is deprecated and not in project deps. "
        "Covered by Phase 3 integration test with a checked-in sample .xls file. "
        "One-exception to 'no skips' rule — documented in 01-02-SUMMARY.md."
    )


# ---------------------------------------------------------------------------
# ODS reading (INP-03)
# ---------------------------------------------------------------------------


def test_read_ods(tmp_path: pathlib.Path) -> None:
    """INP-03: read_ods reads an ODS file produced by pandas + odfpy."""
    ods_path = tmp_path / "data.ods"
    df = pd.DataFrame(
        {
            "nº mec.": ["f6688", "f1234", "f9001", "d5500", "b7700"],
            "nome": [
                "João Silva Teste",
                "Maria Costa Exemplo",
                "Ana Pereira Sintetica",
                "Carlos Oliveira Teste",
                "Rui Ferreira Exemplo",
            ],
        }
    )
    df.to_excel(ods_path, engine="odf", index=False)

    result = read_ods(ods_path)
    assert isinstance(result, ReadResult)
    assert isinstance(result.rows, list)
    # Header row is prepended, so total = 1 header + 5 data rows
    assert len(result.rows) >= 5
    assert result.raw_bytes_sample is None
    assert result.skipped_trailing_empty >= 0


def test_read_ods_permission_error(tmp_path: pathlib.Path) -> None:
    """INP-13: PermissionError from ODS read is wrapped as FileAccessError."""
    ods_path = tmp_path / "locked.ods"
    ods_path.write_bytes(b"")

    with (
        patch("eleitorum.core.readers.pd.read_excel", side_effect=PermissionError("locked")),
        pytest.raises(FileAccessError),
    ):
        read_ods(ods_path)


# ---------------------------------------------------------------------------
# CSV reading (INP-04, INP-05, INP-07)
# ---------------------------------------------------------------------------


def test_read_csv_utf8_bom(tmp_path: pathlib.Path) -> None:
    """INP-04, INP-07: read CSV with UTF-8-BOM encoding via make_simple_caderno."""
    csv_path = make_simple_caderno(tmp_path / "caderno.csv")
    result = read_csv_like(csv_path, delimiter=";", encoding="utf-8-sig")
    assert isinstance(result, ReadResult)
    assert isinstance(result.rows, list)
    # 1 header + 20 data rows
    assert len(result.rows) == 21
    assert result.sheet_name is None
    # raw_bytes_sample must be populated for CSV
    assert result.raw_bytes_sample is not None
    assert len(result.raw_bytes_sample) > 0


def test_read_csv_utf8_no_bom(tmp_path: pathlib.Path) -> None:
    """INP-07: read CSV without BOM using utf-8 encoding."""
    csv_path = tmp_path / "no_bom.csv"
    # Use write_bytes to avoid platform-specific newline translation
    csv_path.write_bytes(
        "nº mec.;nome\r\nf6688;João Silva Teste\r\nf1234;Maria Costa Exemplo\r\n".encode()
    )
    result = read_csv_like(csv_path, delimiter=";", encoding="utf-8")
    assert len(result.rows) == 3  # 1 header + 2 data rows
    assert result.raw_bytes_sample is not None


def test_read_csv_cp1252(tmp_path: pathlib.Path) -> None:
    """INP-07: read CSV encoded as CP1252 by specifying encoding='cp1252'."""
    csv_path = tmp_path / "cp1252.csv"
    csv_path.write_bytes("a;b\r\né;ç\r\n".encode("cp1252"))
    result = read_csv_like(csv_path, delimiter=";", encoding="cp1252")
    assert len(result.rows) == 2  # 1 header + 1 data row
    assert result.rows[1][0] == "é"
    assert result.rows[1][1] == "ç"


def test_read_tsv(tmp_path: pathlib.Path) -> None:
    """INP-05: read a TSV file using tab delimiter."""
    tsv_path = tmp_path / "data.tsv"
    # Use write_bytes to avoid platform-specific newline translation
    tsv_path.write_bytes(
        "nº mec.\tnome\r\nf6688\tJoão Silva Teste\r\nf1234\tMaria Costa Exemplo\r\n".encode()
    )
    result = read_csv_like(tsv_path, delimiter="\t", encoding="utf-8")
    assert len(result.rows) == 3  # 1 header + 2 data rows
    assert result.rows[1][0] == "f6688"
    assert result.rows[1][1] == "João Silva Teste"


def test_csv_raw_bytes_sample_contains_64kb_or_full_file(tmp_path: pathlib.Path) -> None:
    """INP-07: raw_bytes_sample is first 64KB of file (or full file if smaller)."""
    csv_path = make_simple_caderno(tmp_path / "caderno.csv")
    result = read_csv_like(csv_path, delimiter=";", encoding="utf-8-sig")
    file_size = csv_path.stat().st_size
    assert result.raw_bytes_sample is not None
    assert len(result.raw_bytes_sample) == min(file_size, 64 * 1024)


def test_csv_permission_error_raises_file_access_error(tmp_path: pathlib.Path) -> None:
    """INP-13: PermissionError on CSV open is wrapped as FileAccessError."""
    csv_path = tmp_path / "locked.csv"
    csv_path.write_text("a;b\r\n", encoding="utf-8")

    with (
        patch("builtins.open", side_effect=PermissionError("locked")),
        pytest.raises(FileAccessError) as exc_info,
    ):
        read_csv_like(csv_path, delimiter=";")

    assert "aberto" in exc_info.value.message_pt or "abrir" in exc_info.value.message_pt


# ---------------------------------------------------------------------------
# read_input dispatch (INP-01, INP-06)
# ---------------------------------------------------------------------------


def test_unsupported_extension_raises(tmp_path: pathlib.Path) -> None:
    """INP-06: read_input with unsupported extension raises UnsupportedFormatError."""
    docx_path = tmp_path / "x.docx"
    docx_path.write_text("not a real docx")
    with pytest.raises(UnsupportedFormatError) as exc_info:
        read_input(docx_path)
    assert ".docx" in exc_info.value.message_pt


def test_unsupported_extension_no_builtin_exception(tmp_path: pathlib.Path) -> None:
    """INP-06: unsupported extension must not leak a raw Python exception."""
    bad_path = tmp_path / "data.pdf"
    bad_path.write_bytes(b"")
    with pytest.raises(UnsupportedFormatError):
        read_input(bad_path)


def test_read_input_dispatches_by_extension(tmp_path: pathlib.Path) -> None:
    """Dispatch table: .xlsx -> read_xlsx, .csv -> read_csv_like, .docx -> error."""
    xlsx_path = make_multi_sheet_xlsx(tmp_path / "test.xlsx")

    with patch("eleitorum.core.readers.read_xlsx", wraps=read_xlsx) as mock_xlsx:
        read_input(xlsx_path)
        assert mock_xlsx.call_count == 1

    # CSV dispatch
    csv_path = make_simple_caderno(tmp_path / "caderno.csv")
    with patch("eleitorum.core.readers.read_csv_like", wraps=read_csv_like) as mock_csv:
        read_input(csv_path)
        assert mock_csv.call_count == 1

    # Unsupported extension
    bad_path = tmp_path / "data.docx"
    bad_path.write_bytes(b"")
    with pytest.raises(UnsupportedFormatError):
        read_input(bad_path)


def test_read_input_xlsm_dispatches_to_xlsx(tmp_path: pathlib.Path) -> None:
    """read_input with .xlsm calls read_xlsx (same engine as .xlsx)."""
    xlsm_path = tmp_path / "test.xlsm"
    # Create a minimal valid XLSX saved with .xlsm extension
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Sheet1")
    ws.append(["col_a", "col_b"])
    ws.append(["val1", "val2"])
    wb.save(xlsm_path)
    wb.close()

    with patch("eleitorum.core.readers.read_xlsx", wraps=read_xlsx) as mock_xlsx:
        read_input(xlsm_path)
        assert mock_xlsx.call_count == 1


def test_read_input_tsv_dispatches_with_tab_delimiter(tmp_path: pathlib.Path) -> None:
    """read_input with .tsv calls read_csv_like with tab delimiter."""
    tsv_path = tmp_path / "data.tsv"
    tsv_path.write_bytes(b"a\tb\r\n1\t2\r\n")

    with patch("eleitorum.core.readers.read_csv_like", wraps=read_csv_like) as mock_csv:
        read_input(tsv_path)
        call_kwargs = mock_csv.call_args
        assert call_kwargs is not None
        # delimiter must be tab
        args, kwargs = call_kwargs
        delimiter_val = kwargs.get("delimiter") or (args[1] if len(args) > 1 else None)
        assert delimiter_val == "\t"


# ---------------------------------------------------------------------------
# Multi-sheet metadata (INP-10, INP-11)
# ---------------------------------------------------------------------------


def test_multi_sheet_xlsx_returns_sheet_names_and_counts(tmp_path: pathlib.Path) -> None:
    """INP-10: list_sheets returns SheetInfo for each sheet in the correct order."""
    xlsx_path = make_multi_sheet_xlsx(tmp_path / "multi.xlsx")
    infos = list_sheets(xlsx_path)
    assert isinstance(infos, list)
    assert len(infos) == 3
    assert [i.name for i in infos] == ["Docentes", "PTAG", "Alunos"]
    assert all(isinstance(i, SheetInfo) for i in infos)


def test_empty_sheet_flagged(tmp_path: pathlib.Path) -> None:
    """INP-11: Alunos sheet (header only, no data rows) is flagged as is_empty=True."""
    xlsx_path = make_multi_sheet_xlsx(tmp_path / "multi.xlsx")
    infos = list_sheets(xlsx_path)
    alunos = next(i for i in infos if i.name == "Alunos")
    assert alunos.is_empty is True


def test_non_empty_sheets_not_flagged(tmp_path: pathlib.Path) -> None:
    """INP-11: Sheets with data rows must have is_empty=False."""
    xlsx_path = make_multi_sheet_xlsx(tmp_path / "multi.xlsx")
    infos = list_sheets(xlsx_path)
    docentes = next(i for i in infos if i.name == "Docentes")
    ptag = next(i for i in infos if i.name == "PTAG")
    assert docentes.is_empty is False
    assert ptag.is_empty is False


def test_list_sheets_csv_returns_empty(tmp_path: pathlib.Path) -> None:
    """list_sheets returns [] for CSV/TSV files (no sheet concept)."""
    csv_path = make_simple_caderno(tmp_path / "caderno.csv")
    result = list_sheets(csv_path)
    assert result == []


def test_list_sheets_tsv_returns_empty(tmp_path: pathlib.Path) -> None:
    """list_sheets returns [] for TSV files."""
    tsv_path = tmp_path / "data.tsv"
    tsv_path.write_bytes(b"a\tb\r\n")
    result = list_sheets(tsv_path)
    assert result == []


# ---------------------------------------------------------------------------
# Trailing empty row stripping (INP-12)
# ---------------------------------------------------------------------------


def test_skip_trailing_empty_rows_logged_count(tmp_path: pathlib.Path) -> None:
    """INP-12: XLSX with trailing all-None rows strips them and reports count."""
    xlsx_path = tmp_path / "trailing.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Sheet1")
    ws.title = "Sheet1"
    # 5 data rows
    ws.append(["header_a", "header_b"])
    ws.append(["f6688", "João Silva Teste"])
    ws.append(["f1234", "Maria Costa Exemplo"])
    ws.append(["f9001", "Ana Pereira Sintetica"])
    ws.append(["d5500", "Carlos Oliveira Teste"])
    ws.append(["b7700", "Marta Rodrigues Teste"])
    # 3 trailing empty rows (all None)
    ws.append([None, None])
    ws.append([None, None])
    ws.append([None, None])
    wb.save(xlsx_path)
    wb.close()

    result = read_xlsx(xlsx_path)
    assert result.skipped_trailing_empty == 3
    # 6 non-empty rows (1 header + 5 data)
    assert len(result.rows) == 6


def test_strip_trailing_empty_internal() -> None:
    """_strip_trailing_empty correctly identifies and strips all-empty trailing rows."""
    rows: list[tuple[Any, ...]] = [
        ("a", "b"),
        ("c", "d"),
        (None, None),
        ("", "  "),
        (None, ""),
    ]
    stripped, count = _strip_trailing_empty(rows)
    assert count == 3
    assert len(stripped) == 2
    assert stripped[0] == ("a", "b")
    assert stripped[1] == ("c", "d")


def test_strip_trailing_empty_empty_input() -> None:
    """_strip_trailing_empty handles empty list gracefully."""
    stripped, count = _strip_trailing_empty([])
    assert stripped == []
    assert count == 0


def test_strip_trailing_empty_all_empty() -> None:
    """_strip_trailing_empty strips all rows if all are empty."""
    rows: list[tuple[Any, ...]] = [(None, None), ("", ""), (None, "")]
    stripped, count = _strip_trailing_empty(rows)
    assert stripped == []
    assert count == 3


def test_strip_trailing_empty_no_trailing() -> None:
    """_strip_trailing_empty returns all rows unchanged when none are trailing-empty."""
    rows: list[tuple[Any, ...]] = [("a", "b"), ("c", "d"), ("e", "f")]
    stripped, count = _strip_trailing_empty(rows)
    assert stripped == rows
    assert count == 0


# ---------------------------------------------------------------------------
# Additional coverage tests (INP-02 mock, ODS sheet_name, list_sheets ODS)
# ---------------------------------------------------------------------------


def test_read_xls_permission_error(tmp_path: pathlib.Path) -> None:
    """INP-13: PermissionError from xlrd is wrapped as FileAccessError."""
    xls_path = tmp_path / "locked.xls"
    xls_path.write_bytes(b"")  # file exists but xlrd cannot open it

    import xlrd as xlrd_module

    with (
        patch.object(xlrd_module, "open_workbook", side_effect=PermissionError("locked")),
        pytest.raises(FileAccessError) as exc_info,
    ):
        from eleitorum.core.readers import read_xls

        read_xls(xls_path)

    assert "aberto" in exc_info.value.message_pt or "abrir" in exc_info.value.message_pt


def test_read_xls_file_not_found(tmp_path: pathlib.Path) -> None:
    """INP-13: FileNotFoundError from xlrd is wrapped as FileAccessError."""
    missing_path = tmp_path / "missing.xls"

    import xlrd as xlrd_module

    with (
        patch.object(xlrd_module, "open_workbook", side_effect=FileNotFoundError("not found")),
        pytest.raises(FileAccessError),
    ):
        from eleitorum.core.readers import read_xls

        read_xls(missing_path)


def test_read_ods_with_sheet_name(tmp_path: pathlib.Path) -> None:
    """INP-03: read_ods with explicit sheet_name sets chosen_sheet correctly."""
    ods_path = tmp_path / "named.ods"
    df = pd.DataFrame({"col_a": [1, 2], "col_b": ["x", "y"]})
    df.to_excel(ods_path, engine="odf", index=False)

    result = read_ods(ods_path, sheet_name="Sheet1")
    assert result.sheet_name == "Sheet1"
    assert len(result.rows) >= 1


def test_read_ods_file_not_found(tmp_path: pathlib.Path) -> None:
    """INP-13: FileNotFoundError from ODS read is wrapped as FileAccessError."""
    missing_path = tmp_path / "missing.ods"
    with pytest.raises(FileAccessError):
        read_ods(missing_path)


def test_list_sheets_ods(tmp_path: pathlib.Path) -> None:
    """INP-10: list_sheets works for ODS files."""
    ods_path = tmp_path / "data.ods"
    df = pd.DataFrame({"nº mec.": ["f6688", "f1234"], "nome": ["Alice Teste", "Bob Exemplo"]})
    df.to_excel(ods_path, engine="odf", index=False)

    infos = list_sheets(ods_path)
    assert isinstance(infos, list)
    assert len(infos) >= 1
    assert all(isinstance(i, SheetInfo) for i in infos)


def test_csv_file_not_found_raises_file_access_error(tmp_path: pathlib.Path) -> None:
    """INP-13: FileNotFoundError for CSV is wrapped as FileAccessError."""
    missing_path = tmp_path / "missing.csv"
    with pytest.raises(FileAccessError):
        read_csv_like(missing_path, delimiter=";")


def test_read_input_ods_dispatches(tmp_path: pathlib.Path) -> None:
    """read_input with .ods calls read_ods."""
    ods_path = tmp_path / "data.ods"
    df = pd.DataFrame({"col_a": [1, 2], "col_b": ["x", "y"]})
    df.to_excel(ods_path, engine="odf", index=False)

    with patch("eleitorum.core.readers.read_ods", wraps=read_ods) as mock_ods:
        read_input(ods_path)
        assert mock_ods.call_count == 1


def test_read_input_xls_dispatches(tmp_path: pathlib.Path) -> None:
    """read_input with .xls calls read_xls."""
    xls_path = tmp_path / "data.xls"
    xls_path.write_bytes(b"")  # file exists; we mock xlrd

    import xlrd as xlrd_module

    mock_sheet = MagicMock()
    mock_sheet.get_rows.return_value = iter([])
    mock_sheet.name = "Sheet1"
    mock_wb = MagicMock()
    mock_wb.sheet_names.return_value = ["Sheet1"]
    mock_wb.sheet_by_index.return_value = mock_sheet
    mock_wb.release_resources.return_value = None

    with patch.object(xlrd_module, "open_workbook", return_value=mock_wb):
        result = read_input(xls_path)
        assert isinstance(result, ReadResult)
